import openpyxl
from docx import Document
from pathlib import Path

# Caminhos dos arquivos e pastas
planilha_path = r"C:\Users\savio\OneDrive\Área de Trabalho\projetos pro github\Locadora_contratos\dados_clientes_contratos.xlsx"
contratos_dir = Path(r"C:\Users\savio\OneDrive\Área de Trabalho\projetos pro github\Locadora_contratos\contratos")

# Cria o diretório para salvar contratos, se não existir
contratos_dir.mkdir(parents=True, exist_ok=True)

# Modelo do contrato
modelo_contrato = """
CONTRATO DE LOCAÇÃO DE VEÍCULO

LOCADORA: [Nome da Locadora], inscrita no CNPJ sob o nº [CNPJ da Locadora], com sede em [Endereço da Locadora].

LOCATÁRIO:
- NOME: {NOME}
- CPF: {CPF}
- ENDEREÇO: {ENDERECO}
- DATA DE NASCIMENTO: {DATA_DE_NASCIMENTO}
- TELEFONE DE CONTATO: {TELEFONE_DE_CONTATO}
- E-MAIL: {EMAIL}
- TEMPO DE LOCAÇÃO: {TEMPO_DE_LOCACAO} dias

CLÁUSULA PRIMEIRA – DO OBJETO DO CONTRATO
A LOCADORA disponibiliza ao LOCATÁRIO o veículo [Descrição do Veículo: marca, modelo, ano, placa], em perfeito estado de uso e conservação, para o período indicado no campo TEMPO DE LOCAÇÃO, conforme os termos e condições deste contrato.

CLÁUSULA SEGUNDA – DAS OBRIGAÇÕES DO LOCATÁRIO
O LOCATÁRIO se compromete a:
1. Utilizar o veículo exclusivamente para fins lícitos e em conformidade com as leis de trânsito;
2. Devolver o veículo nas mesmas condições em que foi recebido;
3. Arcar com os custos de combustíveis, multas, pedágios e outros encargos durante o período de locação.

CLÁUSULA TERCEIRA – DA DEVOLUÇÃO
O veículo deverá ser devolvido à LOCADORA na data e local previamente estabelecidos. O não cumprimento implicará cobrança de taxas adicionais conforme tabela vigente.

CLÁUSULA QUARTA – DO PAGAMENTO
O LOCATÁRIO concorda em pagar o valor acordado pelo tempo de locação indicado no campo TEMPO DE LOCAÇÃO, conforme os termos descritos na proposta comercial anexada.

CLÁUSULA QUINTA – DAS DISPOSIÇÕES GERAIS
Este contrato é firmado entre as partes em conformidade com as leis vigentes no território nacional.

Assinado e datado em [Cidade], [Data].

ASSINATURA DA LOCADORA
_______________________________
[Nome da Locadora]

ASSINATURA DO LOCATÁRIO
_______________________________
{NOME}
"""

# Função para criar um contrato individual
def gerar_contrato(dados_cliente):
    # Cria o documento do Word
    doc = Document()
    contrato_texto = modelo_contrato.format(**dados_cliente)
    doc.add_paragraph(contrato_texto)

    # Define o nome do arquivo
    nome_arquivo = contratos_dir / f"Contrato_{dados_cliente['NOME'].replace(' ', '_')}.docx"

    # Salva o contrato
    doc.save(nome_arquivo)
    print(f"Contrato gerado: {nome_arquivo}")

# Lê a planilha
wb = openpyxl.load_workbook(planilha_path)
ws = wb.active

# Obtém os dados das colunas
colunas = [cell.value for cell in ws[1]]
clientes = [
    dict(zip(colunas, [cell if cell is not None else '' for cell in row]))  # Garante que células vazias sejam tratadas
    for row in ws.iter_rows(min_row=2, values_only=True)
]

# Verifica se todas as colunas necessárias estão presentes
colunas_necessarias = ['NOME', 'CPF', 'ENDERECO', 'DATA_DE_NASCIMENTO', 'TELEFONE_DE_CONTATO', 'EMAIL', 'TEMPO_DE_LOCACAO']
for col in colunas_necessarias:
    if col not in colunas:
        print(f"A coluna '{col}' não foi encontrada na planilha!")
        exit(1)

# Gera contratos para todos os clientes
for cliente in clientes:
    gerar_contrato(cliente)

print(f"Todos os contratos foram gerados na pasta: {contratos_dir}")
